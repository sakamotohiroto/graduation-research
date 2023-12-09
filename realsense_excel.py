import cv2
import numpy as np
import time
import pyrealsense2 as pyrs2
import mediapipe as mp
import openpyxl
from scipy import signal
def main():
# カメラの設定
    rscfg = pyrs2.config()

    # RGB映像用のカメラの設定
    rscfg.enable_stream(pyrs2.stream.color, 640, 480, pyrs2.format.bgr8, 30)

    # 距離データ取得用のカメラの設定
    rscfg.enable_stream(pyrs2.stream.depth, 640, 480, pyrs2.format.z16, 30) #元は640,480

# パイプライン構築. パイプラインが設定されると、新しいフレームを待つループができる。
    cap = pyrs2.pipeline()

# ストリーミング開始　ストリーミングとは、映像を再生すること。
    profile = cap.start(rscfg)

# 距離[m] = depth * depth_scale 
    depth_sensor = profile.get_device().first_depth_sensor()
    depth_scale = depth_sensor.get_depth_scale()
    clipping_distance_in_meters = 1.0 # meter
    clipping_distance = clipping_distance_in_meters / depth_scale

# Alignオブジェクト生成
# rs.align を使用すると、深度フレームと他のフレームの位置合わせを実行できます。
# 「align_to」は、深度フレームを整列させる予定のストリームタイプです。
# この２行は、定義してるだけで、４４行目でこれらの値を使う。
    align_to = pyrs2.stream.color
    align = pyrs2.align(align_to)

    t1 = time.perf_counter() #波形成のために時間とる。t2と一緒に使う。

# mediapipe 事前設定
    mp_drawing = mp.solutions.drawing_utils  # 描画用のインスタンス
    mp_drawing_styles = mp.solutions.drawing_styles
    mp_pose = mp.solutions.pose

    elapsed_time = 0

    wb = openpyxl.load_workbook('./Book1.xlsx')
    ws = wb['Sheet1']
    for row in ws:
        for cell in row:
            cell.value = None

    flag = 0
    time_arry = []
    val_arry = []

    sample_freq = 30
    cutoff_freq1 = 1 #0.05(20秒で1回の呼吸)Hz以上の周波数を除去する。
    filter_order = 2
    cutoff_freq2 = 0.3 #8秒に１回なら0.125
    sos = signal.butter(filter_order, cutoff_freq1, 'lowpass', output='sos', fs=sample_freq)
    sos2 = signal.butter(filter_order, cutoff_freq2, 'lowpass', output='sos', fs=sample_freq)

    with mp_pose.Pose(
        min_detection_confidence=0.5,
        min_tracking_confidence=0.5) as pose:  #poseを使う宣言と設定
        while elapsed_time<=62:
            # フレーム情報を取得
            frames = cap.wait_for_frames()

            # 深度画像とRGB画像の位置合わせ
            aligned_frames = align.process(frames)

            # framesから、フレームデータを取得する。RGBフレームと深度フレームそれぞれ。
            color_frame = aligned_frames.get_color_frame()# Get RGB frame
            depth_frame = aligned_frames.get_depth_frame()# Get depth frame

            #画像の取得に失敗したら飛ばす
            if not depth_frame or not color_frame:
                continue

            # フレームデータを画像データに変換
            color_image = np.asanyarray(color_frame.get_data())# フレームを行列化
            depth_color_frame = pyrs2.colorizer().colorize(depth_frame)
            depth_image = np.asanyarray(depth_color_frame.get_data())


            # 画像の書き換えを不可にする。パフォーマンス向上のため。
            color_image.flags.writeable = False

            # OpenCVとMediaPipeでRGBの並びが違うため、処理前に変換しておく。
            # CV2:BGR → MediaPipe:RGB
            color_image = cv2.cvtColor(color_image, cv2.COLOR_BGR2RGB)


            ht, wt, _ = color_image.shape

            # 推論処理
            results = pose.process(color_image)

            # 座標の合わせ
            pose_points = []
            pose_list = []
            if results.pose_landmarks is not None:
                for i, point in enumerate(results.pose_landmarks.landmark):
                    x = max(1, min(int(point.x * wt), wt-1))
                    y = max(1, min(int(point.y * ht), ht-1))
                    z = int(point.z * wt)
                    pose_points.append([x, y, z, point.visibility])
                pose_list.append(pose_points)

            
             # 前処理の変換を戻しておく。
            color_image.flags.writeable = True
            color_image = cv2.cvtColor(color_image, cv2.COLOR_RGB2BGR)

            #mediapipeのレンダリング
            if len(pose_list)==1:
                poseKeys = pose_list[0]     # sholder_l = pose[11]; sholder_l[0]...x座標，sholder_l[1]...y座標
                for point in poseKeys:
                    cv2.circle(color_image, (point[0], point[1]), 5, [0,255, 0], -1)
                # print("(x,y,z)=", poseKeys[11][0], poseKeys[11][1], poseKeys[11][2])
                # print("(x,y,z)=", poseKeys[12][0], poseKeys[12][1], poseKeys[11][2])

            #四角形の中の深度を平均化して値をだす。
            right_sholder_x = poseKeys[11][0]-30
            left_sholder_x = poseKeys[12][0]+30
            right_sholder_y = poseKeys[11][1]+70
            left_sholder_y = poseKeys[12][1]+30
            # numpyのmean関数は、配列の要素の平均を取ってくれる。
            ave = depth_image[left_sholder_y:right_sholder_y,left_sholder_x:right_sholder_x]
            mean3 = np.mean(ave) #a3の空間内の配列要素の平均を取る
            t2 = time.perf_counter() #波形成のために時間とる。
            elapsed_time = t2-t1
        
            if cv2.waitKey(1) &  0xFF == ord('s'):
                flag = 1
                print("sボタン押した")
            
            if flag == 1:
                time_arry.append(elapsed_time)
                val_arry.append(mean3)



        # 映像の表示と四角形のレンダリング
            # 設定
            depth_colormap = cv2.applyColorMap(cv2.convertScaleAbs(depth_image, alpha=0.03), cv2.COLORMAP_JET)
            #images = np.hstack((bg_removed, depth_colormap))
            images = np.hstack((color_image, depth_colormap)) #depth_imageに変えたら深度カメラを表示できる

            # 四角形を作る線の形成
            cv2.rectangle(images,(right_sholder_x,right_sholder_y), (left_sholder_x, left_sholder_y), color=(0,0,255),thickness= 4)


            # 映像の表示
            cv2.imshow('RealSense', images)


            # q押したら終了
            if cv2.waitKey(1) &  0xFF == ord('q'):
                break

        

        # カメラの解放
        cap.stop()
        cv2.destroyAllWindows()

        val_arry1 = signal.sosfiltfilt(sos, val_arry)
        val_arry2 = signal.sosfiltfilt(sos2, val_arry)
        for i in range(0,len(time_arry)):
            ws.cell(i+1,1,value=time_arry[i])
            ws.cell(i+1,2,value=val_arry1[i]-val_arry2[i])
        wb.save("Book1.xlsx")
        # for i in range(0,len(time_arry)):
        #     print(val_arry[i])

   

# run---------------------------------------------------------------------------------------
if __name__ == '__main__':
  main()
